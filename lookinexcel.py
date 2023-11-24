from openpyxl import Workbook, load_workbook
import re  

wb = load_workbook('example_excel_document.xlsx') #Document exported from google classroom, opened and saved as xlsx
wb.create_sheet("Rearranged")

ws1 = wb.active
ws2 = wb["Rearranged"] #Created sheet named Rearranged

for i in range(4,60):
    a = wb.active[f'A{i}'].value.split(",") # words seperated from comma
    a = list(filter(lambda x:x, a)) # Empty values filtered
    ws2.append(a) #Values added one by one

for column in ["D","E","F","G","H"]: # D means D1,D2,D3... 
    for cell in ws2[column]:
         
        a = cell.value.find('.')
        b = cell.value[:a] #Since numbers are separated by dots such as 100.0, they are not perceived as numbers. This gives the part before the dot
        cell.value = b
        cell.number_format = '0.00' #Changes format of cell from str to number
        cell.value = int(cell.value) #Written values should also be converted to int in python as well


for row in range( 1,len(ws2["A"])+1 ):
    ws2[f"I{row}"].value = f"=AVERAGE(D{row}:H{row})" #Last column will be average

wb.save("Report.xlsx")

"""
#To look grades on terminal

grades = {}
say = 1
for surname in ws2['A']:
    grades[surname.value] = ws2[f'H{say}'].value  #<<<  D E F G H
    say+=1
 
print(grades)
"""
